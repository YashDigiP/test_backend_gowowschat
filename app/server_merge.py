from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os, re, traceback, hashlib
from werkzeug.utils import secure_filename
from sqlalchemy import create_engine, text

from langchain_ollama import ChatOllama
from langchain_community.vectorstores import FAISS
from langchain_ollama import OllamaEmbeddings
from langchain_community.document_loaders import PyPDFLoader
from langchain.chains import RetrievalQA
KNOWN_TABLES = {"tasks", "client", "client_audit", "tasks_audit"}

import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from io import BytesIO
from fpdf import FPDF
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl.chart import BarChart, Reference



app = Flask(__name__)
CORS(app)
@app.route("/ask-db", methods=["POST"]) 
def ask_db():
    data = request.get_json()
    question = data.get("message")
    if not question:
        return jsonify({"error": "No question provided"}), 400

    from langchain_ollama import ChatOllama
    import re, traceback

    system_prompt = """
    You are an expert PostgreSQL assistant. Answer only with SQL.
    - Use only valid column names.
    - Never join tables unless both are explicitly needed.
    - Never join audit tables or perform INSERT, UPDATE, DELETE.
    - Do not return explanations or markdown — just SQL.
    """

    prompt = f"{system_prompt}\n\nQuestion: {question}"

    try:
        llm = ChatOllama(model="mistral")
        raw_response = llm.invoke(prompt).content.strip()
        print("[Generated SQL]:", raw_response)
# Remove markdown fences only — preserve full SQL including CTEs
        cleaned_sql = raw_response.strip()
        if cleaned_sql.lower().startswith("```sql"):
            cleaned_sql = cleaned_sql[6:]
        if cleaned_sql.endswith("```"):
            cleaned_sql = cleaned_sql[:-3]
        cleaned_sql = cleaned_sql.strip()

        print("[Cleaned SQL]:", cleaned_sql)

        # ✅ Check for valid table names
        def get_valid_table_names():
            with engine.connect() as conn:
                result = conn.execute(text("""
                    SELECT table_name FROM information_schema.tables
                    WHERE table_schema = 'public'
                """))
                return {row[0].lower() for row in result.fetchall()}

        # ✅ Extract real tables only (skip those defined in WITH clauses)
        def get_cte_names(sql):
            cte_matches = re.findall(r'\bwith\s+([a-zA-Z0-9_]+)\s+as\s*\(', sql, re.IGNORECASE)
            return set(cte_matches)

        valid_tables = get_valid_table_names()
        cte_names = get_cte_names(cleaned_sql)

        tables_in_query = set(re.findall(r"\b(?:from|join)\s+(\w+)", cleaned_sql, re.IGNORECASE))
        for table in tables_in_query:
            if table.lower() not in valid_tables and table.lower() not in cte_names:
                return jsonify({ "error": f"❌ Table '{table}' not found. Available: {', '.join(valid_tables)}" }), 400
        statements = [s.strip() for s in cleaned_sql.split(";") if s.strip()]
        response_payload = {}

        for stmt in statements:
            if not stmt.lower().startswith("select"):
                continue

            with engine.connect() as conn:
                try:
                    result = conn.execute(text(stmt))
                    rows = [
                        {key: value for key, value in zip(result.keys(), row)}
                        for row in result.fetchall()
                    ]
                    response_payload["rows"] = rows

                    # Detect chart type
                    chart_type = "bar"
                    question_lower = question.lower()
                    if "pie" in question_lower:
                        chart_type = "pie"
                    elif "donut" in question_lower or "doughnut" in question_lower:
                        chart_type = "doughnut"
                    elif "bar" in question_lower:
                        chart_type = "bar"

                    # ✅ Build chart + summary if valid 2-col data
                    if len(rows) > 0 and len(rows[0]) == 2:
                        keys = list(rows[0].keys())
                        x_key, y_key = keys[0], keys[1]
                        x_data = [row[x_key] for row in rows]
                        y_data = [row[y_key] for row in rows]

                        if all(isinstance(val, (int, float)) for val in y_data):
                            response_payload["chart"] = {
                                "type": chart_type,
                                "x": x_data,
                                "y": y_data,
                                "x_label": x_key,
                                "y_label": y_key
                            }
                            # ✅ Add a simple summary
                            total = sum(y_data)
                            top_label = x_data[y_data.index(max(y_data))]
                            summary = f"📊 Total: {total}. Highest: '{top_label}' with {max(y_data)}."
                            response_payload["summary"] = summary
                            print("[Chart Info]:", response_payload["chart"])

                except Exception as e:
                    print("[Query Error]:", e)
                    fallback_sql = "SELECT * FROM tasks"
                    print("[Fallback SQL]:", fallback_sql)

                    fallback_result = conn.execute(text(fallback_sql))
                    fallback_rows = [
                        {key: value for key, value in zip(fallback_result.keys(), row)}
                        for row in fallback_result.fetchall()
                    ]
                    response_payload = {
                        "rows": fallback_rows,
                        "warning": f"Original query failed: {str(e)}. Showing fallback from tasks."
                    }

        print("[Ask-DB Response Payload]:", response_payload)
        return jsonify(response_payload)

    except Exception as e:
        print("[Traceback]:", traceback.format_exc())
        return jsonify({"error": str(e)}), 500


DB_URI = "postgresql://postgres:admin@localhost:5432/mfdb"
engine = create_engine(DB_URI)

UPLOAD_FOLDER = "uploaded_pdfs"
EXPORT_FOLDER = "exports"
os.makedirs(EXPORT_FOLDER, exist_ok=True)
KB_ROOT = "KB"
VECTOR_STORE_DIR = "vector_stores"
embedding_model = OllamaEmbeddings(model="nomic-embed-text")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
pdf_text_store = {}
active_tasks = {}

@app.route("/upload-pdf", methods=["POST"])
def upload_pdf():
    if "pdf" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["pdf"]
    if file.filename == "":
        return jsonify({"error": "Empty filename"}), 400
    filename = secure_filename(file.filename)
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)
    try:
        import PyPDF2
        with open(filepath, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            pdf_text_store["latest"] = text
        return jsonify({"message": f"✅ PDF '{filename}' uploaded and processed."})
    except Exception as e:
        return jsonify({"error": f"❌ Failed to read PDF: {str(e)}"}), 500

@app.route("/chat", methods=["POST"])
def chat():
    data = request.get_json()
    message = data.get("message")
    task_id = data.get("task_id", "default")
    if not message:
        return jsonify({"error": "No message provided"}), 400
    active_tasks[task_id] = True
    try:
        llm = ChatOllama(model="mistral")
        response = llm.invoke(message).content.strip()
        if not active_tasks.get(task_id, True):
            return jsonify({"error": "⛔ Stopped by user."})
        return jsonify({"response": response})
    except Exception as e:
        print(e)
        return jsonify({"error": str(e)}), 500
    finally:
        active_tasks.pop(task_id, None)

@app.route("/stop", methods=["POST"])
def stop():
    task_id = request.args.get("task_id", "default")
    active_tasks[task_id] = False
    return jsonify({"message": f"Stopped task {task_id}"}), 200

@app.route("/ask-pdf", methods=["POST"])
def ask_pdf():
    data = request.get_json()
    question = data.get("message")
    if not question:
        return jsonify({"error": "No question provided"}), 400
    if "latest" not in pdf_text_store:
        return jsonify({"error": "No PDF uploaded yet"}), 400
    try:
        llm = ChatOllama(model="mistral")
        prompt = f"Answer the question based on this PDF:\n\n{pdf_text_store['latest'][:4000]}\n\nQuestion: {question}"
        response = llm.invoke(prompt).content.strip()
        return jsonify({"response": response})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/list-kb-folders")
def list_kb_folders():
    try:
        folders = [f for f in os.listdir(KB_ROOT) if os.path.isdir(os.path.join(KB_ROOT, f))]
        return jsonify(folders)
    except Exception:
        return jsonify([])

@app.route("/list-kb-folder")
def list_kb_folder():
    folder = request.args.get("folder", "")
    full_path = os.path.join(KB_ROOT, folder)
    try:
        items = []
        for name in os.listdir(full_path):
            item_path = os.path.join(full_path, name)
            if os.path.isdir(item_path):
                pdf_count = sum(1 for f in os.listdir(item_path) if f.lower().endswith(".pdf"))
                items.append({"is_folder": True, "name": name, "pdf_count": pdf_count})
            elif name.lower().endswith(".pdf"):
                items.append({"is_folder": False, "name": name})
        return jsonify(items)
    except Exception:
        return jsonify([])

@app.route("/ask-kb", methods=["POST"])
def ask_kb():
    print("🔔 /ask-kb endpoint hit!")
    data = request.get_json()
    print("📥 Payload:", data)
    question = data.get("message")
    pdf_path = data.get("path")
    task_id = data.get("task_id", "default")

    if not question or not pdf_path:
        return jsonify({"error": "Missing question or PDF path"}), 400

    try:
        full_pdf_path = os.path.join("KB", pdf_path).replace("\\", "/")
        hashed = hashlib.sha256(full_pdf_path.encode()).hexdigest()
        vs_path = os.path.join(VECTOR_STORE_DIR, hashed)
        print("[KB] Received path:", pdf_path)
        print("[KB] Full path for hashing:", full_pdf_path)
        print("[KB] Expected vector store:", vs_path)

        if not os.path.exists(vs_path):
            return jsonify({"error": "Vector store not found for this PDF. Please re-embed it."}), 400

        vectordb = FAISS.load_local(vs_path, embedding_model, allow_dangerous_deserialization=True)
        retriever = vectordb.as_retriever()
        llm = ChatOllama(model="mistral")
        qa = RetrievalQA.from_chain_type(llm=llm, retriever=retriever)

        active_tasks[task_id] = True
        response = qa.invoke({"query": question})["result"]
        if not active_tasks.get(task_id, True):
            return jsonify({"error": "⛔ Stopped by user."})
        return jsonify({"response": response})
    except Exception as e:
        print("[Ask-KB Error]", e)
        return jsonify({"error": str(e)}), 500
    finally:
        active_tasks.pop(task_id, None)

def get_all_column_names():
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT table_name, column_name
            FROM information_schema.columns
            WHERE table_schema = 'public'
        """))
        table_cols = {}
        for table, col in result.fetchall():
            table_cols.setdefault(table, set()).add(col)
        return table_cols

@app.route("/download/<filename>")
def download_file(filename):
    path = os.path.join("exports", filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return jsonify({"error": "File not found"}), 404

@app.route("/export-excel", methods=["POST"])
def export_excel():
    print("[🔁] Excel export route called")
    data = request.get_json()
    rows = data.get("rows", [])
    chart_data = data.get("chart", None)

    if not rows:
        print("[⚠️] No data to export")
        return jsonify({"error": "No data to export"}), 400

    df = pd.DataFrame(rows)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        print("[📊] Creating workbook")
        df.to_excel(writer, sheet_name="Report", index=False, startrow=0)
        ws = writer.sheets["Report"]

        if chart_data:
            print("[📈] Chart data found, adding chart")
            from openpyxl.chart import BarChart, Reference

            chart = BarChart()
            chart.title = "Chart"
            chart.x_axis.title = chart_data["x_label"]
            chart.y_axis.title = chart_data["y_label"]

            rows_count = len(df)
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=rows_count + 1)
            cat_ref = Reference(ws, min_col=1, min_row=2, max_row=rows_count + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)

            chart_start_row = rows_count + 3
            ws.add_chart(chart, f"A{chart_start_row}")
        else:
            print("[⚠️] No chart data provided or incomplete")

    output.seek(0)
    print("[✅] Excel saved to memory")

    return send_file(output,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name="report.xlsx")


# PDF Export Route
@app.route("/export-pdf", methods=["POST"])
def export_pdf():
    data = request.get_json()
    rows = data.get("rows", [])
    chart_data = data.get("chart", None)

    if not rows:
        return jsonify({"error": "No data to export"}), 400

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12, style="B")
    pdf.cell(200, 10, txt="Database Query Results", ln=True, align='C')

    # Table
    pdf.set_font("Arial", size=10)
    columns = rows[0].keys()
    for col in columns:
        pdf.cell(40, 10, str(col), border=1, align='C')
    pdf.ln()

    for row in rows:
        for col in columns:
            pdf.cell(40, 10, str(row[col]), border=1, align='C')
        pdf.ln()

    # Chart
    if chart_data:
        try:
            fig, ax = plt.subplots()

            x = chart_data["x"]
            y = chart_data["y"]
            x_label = chart_data["x_label"]
            y_label = chart_data["y_label"]
            colors = chart_data.get("colors", None)

            if colors and len(colors) == len(x):
                ax.bar(x, y, color=colors)
            else:
                ax.bar(x, y)

            ax.set_xlabel(x_label)
            ax.set_ylabel(y_label)
            ax.set_title("Chart")

            img_stream = BytesIO()
            plt.tight_layout()
            plt.savefig(img_stream, format="png")
            plt.close(fig)
            img_stream.seek(0)

            # Save chart to temporary image file
            chart_filename = "temp_chart.png"
            with open(chart_filename, "wb") as f:
                f.write(img_stream.read())

            # Insert into PDF
            pdf.image(chart_filename, x=10, y=pdf.get_y(), w=180)

            # Clean up temp file
            os.remove(chart_filename)

        except Exception as e:
            print(f"[⚠️ Chart Error]: {e}")

    # Return PDF
    pdf_bytes = pdf.output(dest="S").encode("latin1")
    pdf_stream = BytesIO(pdf_bytes)
    pdf_stream.seek(0)

    return send_file(
        pdf_stream,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="report.pdf"
    )





@app.route("/download-pdf", methods=["POST"])
def download_pdf_direct():
    path = os.path.join(EXPORT_FOLDER, "report.pdf")
    if os.path.exists(path):
        return send_file(path, mimetype="application/pdf", as_attachment=True)
    return jsonify({"error": "PDF not found"}), 404

@app.route("/download-excel", methods=["POST"])
def download_excel_direct():
    path = os.path.join(EXPORT_FOLDER, "report.xlsx")
    if os.path.exists(path):
        return send_file(path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True)
    return jsonify({"error": "Excel not found"}), 404

@app.route("/generate-prompts", methods=["POST"])
def generate_prompts():
    data = request.get_json()
    saathi_reply = data.get("text", "")
    if not saathi_reply:
        return jsonify({"prompts": []})

    from langchain_ollama import ChatOllama

    try:
        llm = ChatOllama(model="mistral")
        prompt = f"Based on this response, suggest 3 relevant follow-up user questions as a list:\n\n'{saathi_reply}'"
        result = llm.invoke(prompt).content.strip()
        suggestions = [line.strip("-• ").strip() for line in result.splitlines() if line.strip()]
        print("[Suggested Prompts]:", suggestions[:3])  # ✅ Show in terminal
        return jsonify({"prompts": suggestions[:3]})
    
    except Exception as e:
        print("[/generate-prompts Error]:", str(e))
        return jsonify({"prompts": [], "error": str(e)}), 500







@app.route("/embed-site", methods=["POST"])
def embed_site():
    from langchain_community.document_loaders.recursive_url_loader import RecursiveUrlLoader
    import hashlib

    data = request.get_json()
    url = data.get("url", "").strip()
    if not url:
        return jsonify({"error": "No URL provided"}), 400

    try:
        print("[Ask Web] Crawling:", url)
        hashed = hashlib.sha256(url.encode()).hexdigest()
        vs_path = os.path.join(VECTOR_STORE_DIR, f"web_{hashed}")
        if os.path.exists(vs_path):
            return jsonify({"message": "✅ Already embedded", "id": hashed})

        # Fix: extractor must be a callable, not a string
        loader = RecursiveUrlLoader(url=url, max_depth=2, extractor=lambda x: x)
        docs = loader.load()

        vectordb = FAISS.from_documents(docs, embedding_model)
        vectordb.save_local(vs_path)
        return jsonify({"message": f"✅ Embedded site: {url}", "id": hashed})

    except Exception as e:
        print("[Ask Web Error]:", str(e))
        return jsonify({"error": str(e)}), 500
@app.route("/ask-web", methods=["POST"])
def ask_web():
    from langchain_community.vectorstores import FAISS
    from langchain.chains import RetrievalQA

    data = request.get_json()
    question = data.get("question", "")
    vector_id = data.get("id", "")
    if not question or not vector_id:
        return jsonify({"error": "Missing question or ID"}), 400

    vs_path = os.path.join(VECTOR_STORE_DIR, f"web_{vector_id}")
    if not os.path.exists(vs_path):
        return jsonify({"error": "Vector store not found. Please embed site first."}), 400

    try:
        vectordb = FAISS.load_local(vs_path, embedding_model, allow_dangerous_deserialization=True)
        retriever = vectordb.as_retriever()
        llm = ChatOllama(model="mistral")
        qa = RetrievalQA.from_chain_type(llm=llm, retriever=retriever)
        response = qa.invoke({"query": question})["result"]
        return jsonify({"response": response})

    except Exception as e:
        print("[Ask Web Query Error]:", str(e))
        return jsonify({"error": str(e)}), 500



@app.route("/ask-web-live", methods=["POST"])
def ask_web_live():
    from langchain_community.document_loaders.recursive_url_loader import RecursiveUrlLoader
    from langchain.chains import RetrievalQA
    from langchain_community.vectorstores import FAISS

    data = request.get_json()
    url = data.get("url", "")
    question = data.get("question", "")
    if not url or not question:
        return jsonify({"error": "Missing URL or question"}), 400

    try:
        print(f"[Ask Web Live] Received request for URL: {url}")
        print(f"[Ask Web Live] Question: {question}")
        
        loader = RecursiveUrlLoader(url=url, max_depth=1, extractor=lambda x: x)
        docs = loader.load()
        print(f"[Ask Web Live] Loaded {len(docs)} documents.")

        vectordb = FAISS.from_documents(docs, embedding_model)
        retriever = vectordb.as_retriever()
        llm = ChatOllama(model="mistral")
        qa = RetrievalQA.from_chain_type(llm=llm, retriever=retriever)
        response = qa.invoke({"query": question})["result"]
        print(f"[Ask Web Live] Response: {response}")
        return jsonify({"response": response})

    except Exception as e:
        print("[Ask Web Live Error]:", str(e))
        return jsonify({"error": str(e)}), 500




@app.route("/ask-web-lite", methods=["POST"])
def ask_web_lite():
    from langchain.chains import RetrievalQA
    from langchain_community.vectorstores import FAISS
    from langchain_community.document_loaders import WebBaseLoader

def get_hashed_path(relative_path: str) -> str:
    normalized = os.path.normpath(relative_path).replace("\\", "/")
    return hashlib.sha256(normalized.encode()).hexdigest()


    data = request.get_json()
    url = data.get("url", "")
    question = data.get("question", "")
    if not url or not question:
        return jsonify({"error": "Missing URL or question"}), 400

    try:
        print(f"[Ask Web Lite] Fetching single page: {url}")
        loader = WebBaseLoader(url)
        docs = loader.load()
        print(f"[Ask Web Lite] Loaded {len(docs)} document(s).")

        vectordb = FAISS.from_documents(docs, embedding_model)
        retriever = vectordb.as_retriever()
        llm = ChatOllama(model="mistral")
        qa = RetrievalQA.from_chain_type(llm=llm, retriever=retriever)

        response = qa.invoke({"query": question})["result"]
        print(f"[Ask Web Lite] Response: {response}")
        return jsonify({"response": response})

    except Exception as e:
        print("[Ask Web Lite Error]:", str(e))
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(port=7860, debug=True)